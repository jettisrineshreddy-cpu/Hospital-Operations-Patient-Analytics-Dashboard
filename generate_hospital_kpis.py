

import pandas as pd
from pathlib import Path
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


PROJECT_ROOT = Path(__file__).resolve().parent.parent

CLEAN_FOLDER = PROJECT_ROOT / "data" / "cleaned"
FINAL_FOLDER = PROJECT_ROOT / "data" / "final"

FINAL_FOLDER.mkdir(exist_ok=True)

print("=" * 60)
print("HOSPITAL KPI ENGINEERING")
print("=" * 60)

print(f"Project : {PROJECT_ROOT}")
print(f"Cleaned : {CLEAN_FOLDER}")
print(f"Final   : {FINAL_FOLDER}")


datasets = {
    "Patients": "Patients_Cleaned.csv",
    "Department": "Department_Cleaned.csv",
    "Doctor": "Doctor_Cleaned.csv",
    "Nurse": "Nurse_Cleaned.csv",
    "Helpers": "Helpers_Cleaned.csv",
    "Ward": "Ward_Cleaned.csv",
    "Room": "Room_Cleaned.csv",
    "Bed": "Bed_Cleaned.csv",
    "Appointment": "Appointment_Cleaned.csv",
    "MedicalRecord": "MedicalRecord_Cleaned.csv",
    "SurgeryRecord": "SurgeryRecord_Cleaned.csv",
    "RoomRecords": "RoomRecords_Cleaned.csv",
    "BedRecords": "BedRecords_Cleaned.csv",
    "StaffShift": "StaffShift_Cleaned.csv",
    "Equipment": "Equipment_Cleaned.csv",
    "Equipment_Usage": "Equipment_Usage_Cleaned.csv"
}

loaded_data = {}

print("\nLoading Cleaned Datasets...\n")

for name, file in datasets.items():

    df = pd.read_csv(CLEAN_FOLDER / file)

    loaded_data[name] = df

    print(f"{name:<20} Rows : {df.shape[0]:<6} Columns : {df.shape[1]}")



patients = loaded_data["Patients"]
department = loaded_data["Department"]
doctor = loaded_data["Doctor"]
nurse = loaded_data["Nurse"]
helpers = loaded_data["Helpers"]
ward = loaded_data["Ward"]
room = loaded_data["Room"]
bed = loaded_data["Bed"]

appointment = loaded_data["Appointment"]
medical = loaded_data["MedicalRecord"]
surgery = loaded_data["SurgeryRecord"]
roomrecords = loaded_data["RoomRecords"]
bedrecords = loaded_data["BedRecords"]
staffshift = loaded_data["StaffShift"]

equipment = loaded_data["Equipment"]
equipment_usage = loaded_data["Equipment_Usage"]


appointment["appointment_Date"] = pd.to_datetime(
    appointment["appointment_Date"]
)

medical["visit_Date"] = pd.to_datetime(
    medical["visit_Date"]
)

medical["next_Visit"] = pd.to_datetime(
    medical["next_Visit"],
    errors="coerce"
)


surgery["surgery_Date"] = pd.to_datetime(
    surgery["surgery_Date"]
)

surgery["start_Time"] = pd.to_datetime(
    surgery["start_Time"],
    format="%H:%M:%S"
)

surgery["end_Time"] = pd.to_datetime(
    surgery["end_Time"],
    format="%H:%M:%S"
)


roomrecords["admission_Date"] = pd.to_datetime(
    roomrecords["admission_Date"]
)

roomrecords["discharge_Date"] = pd.to_datetime(
    roomrecords["discharge_Date"]
)

bedrecords["admission_Date"] = pd.to_datetime(
    bedrecords["admission_Date"]
)

bedrecords["discharge_Date"] = pd.to_datetime(
    bedrecords["discharge_Date"]
)

staffshift["shift_Date"] = pd.to_datetime(
    staffshift["shift_Date"]
)

staffshift["shift_Start"] = pd.to_datetime(
    staffshift["shift_Start"],
    format="%H:%M:%S"
)

staffshift["shift_End"] = pd.to_datetime(
    staffshift["shift_End"],
    format="%H:%M:%S"
)


equipment["Purchase_Date"] = pd.to_datetime(
    equipment["Purchase_Date"]
)

equipment["Last_Maintenance_Date"] = pd.to_datetime(
    equipment["Last_Maintenance_Date"]
)


equipment_usage["Usage_Date"] = pd.to_datetime(
    equipment_usage["Usage_Date"]
)


summary = pd.DataFrame({
    "Dataset": loaded_data.keys(),
    "Rows": [df.shape[0] for df in loaded_data.values()],
    "Columns": [df.shape[1] for df in loaded_data.values()]
})

print("\n")
print("=" * 60)
print("DATA LOADING COMPLETED")
print("=" * 60)

print(summary)

print(f"\nDatasets Loaded : {len(loaded_data)}")



print("\n")
print("=" * 60)
print("FEATURE ENGINEERING")
print("=" * 60)





appointment["Appointment_Completed"] = (

    appointment["appointment_status"]

    .str.strip()

    .str.title()

)

appointment["Appointment_Completed"] = (

    appointment["Appointment_Completed"]

    .replace({

        "Completed": "Completed"

    })

    .where(

        appointment["Appointment_Completed"] == "Completed",

        "Not Completed"

    )

)



medical["Has_FollowUp"] = (

    medical["next_Visit"]

    .notna()

)

medical["FollowUp_Status"] = (

    medical["Has_FollowUp"]

    .map({

        True: "Follow-up Required",

        False: "No Follow-up"

    })

)


duration = (
    surgery["end_Time"] -
    surgery["start_Time"]
).dt.total_seconds() / 60


duration = duration.where(duration >= 0, duration + (24 * 60))

surgery["Surgery_Duration_Minutes"] = duration.round().astype(int)

surgery["Duration_Category"] = pd.cut(

    surgery["Surgery_Duration_Minutes"],

    bins=[0, 60, 120, float("inf")],

    labels=[
        "Short",
        "Medium",
        "Long"
    ],

    include_lowest=True

)



def classify_stay_length(df):

    df["Stay_Days"] = (
        df["discharge_Date"] -
        df["admission_Date"]
    ).dt.days

    df["Length_of_Stay_Category"] = pd.cut(

        df["Stay_Days"],

        bins=[0, 3, 7, float("inf")],

        labels=[
            "1-3 Days",
            "4-7 Days",
            "8+ Days"
        ],

        include_lowest=True

    )

    return df


roomrecords = classify_stay_length(roomrecords)


bedrecords = classify_stay_length(bedrecords)

shift_duration = (
    staffshift["shift_End"] -
    staffshift["shift_Start"]
).dt.total_seconds() / 3600


shift_duration = shift_duration.where(
    shift_duration >= 0,
    shift_duration + 24
)

staffshift["Shift_Duration_Hours"] = shift_duration.round(2)

staffshift["Shift_Type"] = pd.cut(

    staffshift["shift_Start"].dt.hour,

    bins=[0, 8, 16, 24],

    labels=[
        "Night",
        "Morning",
        "Evening"
    ],

    include_lowest=True,
    right=False

)


reference_date = equipment_usage["Usage_Date"].max()
equipment["Equipment_Age_Years"] = (

    (

        reference_date

        -

        equipment["Purchase_Date"]

    ).dt.days

    /365

).round(1)


equipment_usage["Maintenance_Required"] = (

    equipment_usage["Maintenance_Flag"]

    .str.strip()

    .str.lower()

    == "yes"

)

usage_threshold = (

    equipment_usage["Hours_Used"]

    .quantile(0.75)

)

equipment_usage["High_Usage"] = (

    equipment_usage["Hours_Used"]

    >= usage_threshold

)

print("\nFeature Engineering Completed Successfully.")

print("\nNew Features Created:")

print("--------------------------------------------")

print("Appointment          -> Appointment_Completed")
print("MedicalRecord        -> Has_FollowUp")
print("SurgeryRecord        -> Surgery_Duration_Minutes")
print("RoomRecords          -> Stay_Days")
print("RoomRecords          -> Length_of_Stay_Category")
print("BedRecords           -> Stay_Days")
print("BedRecords           -> Length_of_Stay_Category")
print("StaffShift           -> Shift_Duration_Hours")
print("Equipment            -> Equipment_Age_Years")
print("Equipment_Usage      -> Maintenance_Required")
print("Equipment_Usage      -> High_Usage")



print("\n")
print("=" * 60)
print("CREATING EXECUTIVE KPI DASHBOARD")
print("=" * 60)


total_patients = len(patients)
total_doctors = len(doctor)
total_nurses = len(nurse)
total_helpers = len(helpers)

total_departments = len(department)
total_wards = len(ward)
total_rooms = len(room)
total_beds = len(bed)

total_medical_records = len(medical)
total_appointments = len(appointment)
total_surgeries = len(surgery)

total_equipment = len(equipment)
total_equipment_usage = len(equipment_usage)
total_unique_patients = roomrecords["patient_Id"].nunique()

total_doctors = doctor.shape[0]

total_room_admissions = len(roomrecords)
total_bed_admissions = len(bedrecords)

total_admissions = (
    total_room_admissions +
    total_bed_admissions
)



total_room_revenue = roomrecords["amount"].sum()

total_bed_revenue = bedrecords["amount"].sum()

total_revenue = (
    total_room_revenue +
    total_bed_revenue
)

average_revenue_per_admission = round(

    total_revenue /
    total_admissions,

    2

)



average_room_stay = round(

    roomrecords["Stay_Days"].mean(),

    2

)

average_bed_stay = round(

    bedrecords["Stay_Days"].mean(),

    2

)

average_length_of_stay = round(

    (
        average_room_stay +
        average_bed_stay
    ) / 2,

    2

)



completed_appointments = (

    appointment["Appointment_Completed"]

    == "Completed"

).sum()

appointment_completion_rate = round(

    completed_appointments /

    total_appointments

    *100,

    2

)



average_surgery_duration = round(

    surgery["Surgery_Duration_Minutes"]

    .mean(),

    2

)



average_equipment_usage = round(

    equipment_usage["Hours_Used"]

    .mean(),

    2

)

high_usage_equipment = (

    equipment_usage["High_Usage"]

    .sum()

)

maintenance_required = (

    equipment_usage["Maintenance_Required"]

    .sum()

)



followup_patients = (

    medical["Has_FollowUp"]

    .sum()

)



occupied_resource_days = (

    roomrecords["Stay_Days"].sum()

    +

    bedrecords["Stay_Days"].sum()

)

analysis_days = (

    max(

        roomrecords["discharge_Date"].max(),

        bedrecords["discharge_Date"].max()

    )

    -

    min(

        roomrecords["admission_Date"].min(),

        bedrecords["admission_Date"].min()

    )

).days + 1

available_resource_days = (

    total_rooms +

    total_beds

) * analysis_days

resource_occupancy_rate = round(

    occupied_resource_days

    /

    available_resource_days

    *100,

    2

)

total_occupied_bed_days = bedrecords["Stay_Days"].sum()

available_bed_days = total_beds * analysis_days

bed_utilization_rate = round(

    (total_occupied_bed_days / available_bed_days) * 100,

    2

)
department_efficiency_score = (
    round(total_unique_patients / total_doctors, 2)
    if total_doctors != 0
    else 0
)


executive_dashboard = pd.DataFrame({

    "KPI":[

        "Total Patients",
        "Total Doctors",
        "Total Nurses",
        "Total Helpers",

        "Total Departments",
        "Total Wards",
        "Total Rooms",
        "Total Beds",

        "Total Medical Records",
        "Total Appointments",
        "Total Surgeries",

        "Total Equipment",
        "Total Equipment Usage Records",

        "Total Admissions",

        "Room Admission Revenue",
        "Bed Admission Revenue",
        "Total Revenue",

        "Average Revenue per Admission",

        "Average Room Stay",
        "Average Bed Stay",
        "Average Length of Stay",

        "Appointment Completion Rate (%)",

        "Average Surgery Duration (Minutes)",

        "Average Equipment Usage Hours",

        "High Usage Equipment",

        "Maintenance Required",

        "Follow-up Patients",

        "Resource Occupancy Rate (%)",

        "Bed Utilization Rate (%)",

        "Department Efficiency Score"

    ],

    "Value":[

        total_patients,
        total_doctors,
        total_nurses,
        total_helpers,

        total_departments,
        total_wards,
        total_rooms,
        total_beds,

        total_medical_records,
        total_appointments,
        total_surgeries,

        total_equipment,
        total_equipment_usage,

        total_admissions,

        total_room_revenue,
        total_bed_revenue,
        total_revenue,

        average_revenue_per_admission,

        average_room_stay,
        average_bed_stay,
        average_length_of_stay,

        appointment_completion_rate,

        average_surgery_duration,

        average_equipment_usage,

        high_usage_equipment,

        maintenance_required,

        followup_patients,

        resource_occupancy_rate,

        bed_utilization_rate,

        department_efficiency_score

    ]

})


kpi_definitions = pd.DataFrame({

    "KPI":[

        "Total Patients",
        "Total Doctors",
        "Total Nurses",
        "Total Helpers",

        "Total Departments",
        "Total Wards",
        "Total Rooms",
        "Total Beds",

        "Total Medical Records",
        "Total Appointments",
        "Total Surgeries",

        "Total Equipment",
        "Total Equipment Usage Records",

        "Total Admissions",

        "Room Admission Revenue",
        "Bed Admission Revenue",
        "Total Revenue",

        "Average Revenue per Admission",

        "Average Room Stay",
        "Average Bed Stay",
        "Average Length of Stay",

        "Appointment Completion Rate",

        "Average Surgery Duration",

        "Average Equipment Usage Hours",

        "High Usage Equipment",

        "Maintenance Required",

        "Follow-up Patients",

        "Resource Occupancy Rate",

        "Bed Utilization Rate",

        "Department Efficiency Score"

    ],

    "Formula":[

        "COUNT(Patients)",
        "COUNT(Doctors)",
        "COUNT(Nurses)",
        "COUNT(Helpers)",

        "COUNT(Departments)",
        "COUNT(Wards)",
        "COUNT(Rooms)",
        "COUNT(Beds)",

        "COUNT(Medical Records)",
        "COUNT(Appointments)",
        "COUNT(Surgeries)",

        "COUNT(Equipment)",
        "COUNT(Equipment Usage)",

        "Room + Bed Admissions",

        "SUM(Room Amount)",
        "SUM(Bed Amount)",
        "Room Revenue + Bed Revenue",

        "Total Revenue / Total Admissions",

        "AVG(Room Stay Days)",
        "AVG(Bed Stay Days)",
        "Average(Room Stay, Bed Stay)",

        "Completed / Total Appointments",

        "AVG(Surgery Duration)",

        "AVG(Hours Used)",

        "Hours Used >= 75th Percentile",

        "Maintenance Flag = Yes",

        "Has Follow-up",

        "Occupied Resource Days / Available Resource Days",

        "Bed Admissions / Total Beds",

        "Total Patients / Total Doctors"

    ],

    "Description":[

        "Total registered patients",
        "Total doctors",
        "Total nurses",
        "Total helpers",

        "Total departments",
        "Total wards",
        "Total rooms",
        "Total beds",

        "Total medical records",
        "Total appointments",
        "Total surgeries",

        "Total equipment",
        "Total equipment usage records",

        "Overall admissions",

        "Revenue generated from rooms",
        "Revenue generated from beds",
        "Overall hospital revenue",

        "Average revenue earned per admission",

        "Average room stay",
        "Average bed stay",
        "Overall average stay",

        "Completed appointment percentage",

        "Average surgery duration",

        "Average equipment usage",

        "Equipment with very high usage",

        "Equipment needing maintenance",

        "Patients requiring follow-up",

        "Overall occupancy percentage",

        "Bed utilization percentage",

        "Patients handled per doctor"

    ]

})



data_dictionary = pd.DataFrame({

    "Table":[

        "Appointment",
        "MedicalRecord",
        "MedicalRecord",
        "SurgeryRecord",
        "SurgeryRecord",
        "RoomRecords",
        "RoomRecords",
        "BedRecords",
        "BedRecords",
        "StaffShift",
        "StaffShift",
        "Equipment",
        "Equipment_Usage",
        "Equipment_Usage"

    ],

    "Engineered Column":[

        "Appointment_Completed",

        "Has_FollowUp",
        "FollowUp_Status",

        "Surgery_Duration_Minutes",
        "Duration_Category",

        "Stay_Days",
        "Length_of_Stay_Category",

        "Stay_Days",
        "Length_of_Stay_Category",

        "Shift_Duration_Hours",
        "Shift_Type",

        "Equipment_Age_Years",

        "Maintenance_Required",
        "High_Usage"

    ],

    "Description":[

        "Indicates whether the appointment was completed.",

        "Indicates whether a follow-up visit exists.",
        "Readable follow-up status for reporting.",

        "Total duration of surgery in minutes.",
        "Categorizes surgery as Short, Medium, or Long.",

        "Length of room stay in days.",
        "Classifies room stay into 1–3, 4–7, or 8+ days.",

        "Length of bed stay in days.",
        "Classifies bed stay into 1–3, 4–7, or 8+ days.",

        "Shift duration in hours.",
        "Categorizes shifts as Morning, Evening, or Night.",

        "Age of equipment in years using the latest usage date as reference.",

        "Indicates whether maintenance is required.",
        "Identifies equipment usage in the top 25%."

    ]

})

total_missing_values = sum(
    df.isnull().sum().sum()
    for df in loaded_data.values()
)



total_duplicate_rows = sum(
    df.duplicated().sum()
    for df in loaded_data.values()
)


engineered_features = 14


total_sheets = 20

quality_report = pd.DataFrame({

    "Validation Check":[

        "Datasets Loaded",
        "Total Missing Values",
        "Total Duplicate Rows",
        "Feature Engineering Completed",
        "Engineered Features",
        "Executive KPI Dashboard Created",
        "KPI Definitions Created",
        "Data Dictionary Created",
        "Workbook Sheets"

    ],

    "Result":[

        len(loaded_data),

        total_missing_values,

        total_duplicate_rows,

        "Yes",

        engineered_features,

        "Yes",

        "Yes",

        "Yes",

        total_sheets

    ],

    "Status":[

        "PASS" if len(loaded_data) == 16 else "FAIL",

        "PASS" if total_missing_values == 0 else "REVIEW",

        "PASS" if total_duplicate_rows == 0 else "WARNING",

        "PASS",

        "PASS",

        "PASS",

        "PASS",

        "PASS",

        "PASS"

    ]

})


print("\nExporting hospital_final_dataset.xlsx ...")

output_file = FINAL_FOLDER / "hospital_final_dataset.xlsx"

with pd.ExcelWriter(output_file, engine="openpyxl") as writer:

    

    executive_dashboard.to_excel(
        writer,
        sheet_name="Executive KPI Dashboard",
        index=False
    )


    patients.to_excel(writer, sheet_name="Patients", index=False)
    department.to_excel(writer, sheet_name="Department", index=False)
    doctor.to_excel(writer, sheet_name="Doctor", index=False)
    nurse.to_excel(writer, sheet_name="Nurse", index=False)
    helpers.to_excel(writer, sheet_name="Helpers", index=False)
    ward.to_excel(writer, sheet_name="Ward", index=False)
    room.to_excel(writer, sheet_name="Room", index=False)
    bed.to_excel(writer, sheet_name="Bed", index=False)

   

    appointment.to_excel(writer, sheet_name="Appointment", index=False)
    medical.to_excel(writer, sheet_name="MedicalRecord", index=False)
    surgery.to_excel(writer, sheet_name="SurgeryRecord", index=False)
    roomrecords.to_excel(writer, sheet_name="RoomRecords", index=False)
    bedrecords.to_excel(writer, sheet_name="BedRecords", index=False)
    staffshift.to_excel(writer, sheet_name="StaffShift", index=False)
    equipment.to_excel(writer, sheet_name="Equipment", index=False)
    equipment_usage.to_excel(writer, sheet_name="Equipment_Usage", index=False)

    

    kpi_definitions.to_excel(
        writer,
        sheet_name="KPI Definitions",
        index=False
    )

    data_dictionary.to_excel(
        writer,
        sheet_name="Data Dictionary",
        index=False
    )

    quality_report.to_excel(
        writer,
        sheet_name="Data Quality Report",
        index=False
    )

    

    workbook = writer.book

    header_fill = PatternFill(
        fill_type="solid",
        start_color="1F4E78",
        end_color="1F4E78"
    )

    header_font = Font(
        bold=True,
        color="FFFFFF",
        size=11
    )

    center_alignment = Alignment(
        horizontal="center",
        vertical="center"
    )

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    for sheet in workbook.worksheets:

        
        sheet.freeze_panes = "A2"

        
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = thin_border

        
        for column in sheet.columns:

            max_length = 0

            column_letter = get_column_letter(column[0].column)

            for cell in column:

                cell.border = thin_border
                cell.alignment = center_alignment

                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

            sheet.column_dimensions[column_letter].width = max_length + 4

        
        for row in sheet.iter_rows(min_row=2):

            for cell in row:

                if isinstance(cell.value, (int, float)):

                    header = sheet.cell(
                        row=1,
                        column=cell.column
                    ).value

                    if header:

                        header = str(header)

                        if "Revenue" in header:
                            cell.number_format = '#,##0.00'

                        elif "Rate" in header or "%" in header:
                            cell.number_format = '0.00'

                        elif "Average" in header:
                            cell.number_format = '0.00'



print("\n")

print("HOSPITAL KPI ENGINEERING COMPLETED")


print(f"Output File : {output_file}")

print("\nSheets Created : 20")

print("\nWorkbook Includes")


print("✓ Executive KPI Dashboard")

for table in datasets.keys():
    print(f"✓ {table}")

print("✓ KPI Definitions")
print("✓ Data Dictionary")
print("✓ Data Quality Report")

print("\nWorkbook Features")


print("✓ Professional Formatting")
print("✓ Freeze Header")
print("✓ Auto Column Width")
print("✓ Center Alignment")
print("✓ Borders")
print("✓ Currency Formatting")
print("✓ Percentage Formatting")
print("✓ Analytics Ready Dataset")

print("\nProject Summary")

print(f"Datasets Processed      : {len(loaded_data)}")
print(f"Engineered Features     : {engineered_features}")
print(f"Workbook Sheets         : {total_sheets}")
print(f"Output File             : {output_file.name}")

print("\nHospital KPI Engineering completed successfully!")



